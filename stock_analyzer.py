#!/usr/bin/env python3
"""
Stock Portfolio Analyzer
========================
Reads stock tickers from a text file, fetches market data via yfinance,
generates an Excel workbook with portfolio analysis, and optionally
calls the Claude API for AI-powered insights.

Tickers file format: TICKER[,SHARES[,BUY_DATE]]  [# inline comment]
  BUY_DATE format: YYYY-MM-DD  (enables P&L columns when provided)

Usage:
    python stock_analyzer.py [tickers_file] [options]

Options:
    --output FILE   Save workbook to this path instead of auto-timestamped name
    --no-ai         Skip AI analysis even if ANTHROPIC_API_KEY is set
    --open          Open the workbook after saving (macOS / Linux / Windows)
    --watch         Re-run automatically whenever the tickers file changes
    --model MODEL   Claude model for AI analysis (default: claude-sonnet-4-6)
    --add ENTRY     Append a ticker line to tickers file and exit
                    e.g.: --add "TSLA,10,2024-03-01"
"""

import sys
import os
import json
import time
import argparse
import subprocess
import re
from datetime import datetime, date as date_type
from pathlib import Path
from typing import Optional

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from dotenv import load_dotenv

try:
    from tqdm import tqdm as _tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    _tqdm = None  # type: ignore

load_dotenv()

# ─── Style constants ──────────────────────────────────────────────────────────
DARK_BLUE     = "1F3864"
MED_BLUE      = "2E75B6"
LIGHT_BLUE    = "D9E1F2"
ALT_ROW       = "F2F2F2"
YELLOW_FILL   = "FFFF00"
GREEN_FILL    = "E2EFDA"
GREEN_TOTAL   = "375623"
RED_CF        = "FFC7CE"
RED_CF_FONT   = "9C0006"
GREEN_CF      = "C6EFCE"
GREEN_CF_FONT = "276221"
ORANGE_FILL   = "FFD966"
DARK_RED      = "C00000"

# ─── Column definitions ───────────────────────────────────────────────────────
# Tuples: (column_letter, header_label, stock_data_key_or_None, number_format, column_width)
#
# Layout:
#   A–D   : Identity  (Ticker, Name, Price, Currency)
#   E–K   : Position  (Buy Date, Buy Price, # Shares, Total Value, Cost Basis, P&L $, P&L %)
#   L–U   : Returns   (1W … 5Y %)
#   V–AE  : Fundamentals (Country, Sector, Industry, Mkt Cap, P/E, 52W Hi/Lo, drawdown, Div, Beta)

COLUMNS = [
    # ── Identity ──
    ("A",  "Ticker",           "ticker",         "@",              10),
    ("B",  "Company Name",     "name",           "@",              28),
    ("C",  "Current Price",    "current_price",  "#,##0.00",       13),
    ("D",  "Currency",         "currency",       "@",               8),
    # ── Position / P&L ──
    ("E",  "Buy Date",         "buy_date",       "yyyy-mm-dd",     12),
    ("F",  "Buy Price",        "buy_price",      "#,##0.00",       12),
    ("G",  "# Shares",         None,             "#,##0.####",     10),
    ("H",  "Total Value",      None,             "#,##0.00",       15),
    ("I",  "Cost Basis",       None,             "#,##0.00",       14),
    ("J",  "P&L $",            None,             "#,##0.00",       13),
    ("K",  "P&L %",            None,             "0.00%",          10),
    # ── Returns ──
    ("L",  "1W%",              "1W%",            "0.00%",           8),
    ("M",  "1M%",              "1M%",            "0.00%",           8),
    ("N",  "3M%",              "3M%",            "0.00%",           8),
    ("O",  "6M%",              "6M%",            "0.00%",           8),
    ("P",  "YTD%",             "YTD%",           "0.00%",           8),
    ("Q",  "1Y%",              "1Y%",            "0.00%",           8),
    ("R",  "2Y%",              "2Y%",            "0.00%",           8),
    ("S",  "3Y%",              "3Y%",            "0.00%",           8),
    ("T",  "4Y%",              "4Y%",            "0.00%",           8),
    ("U",  "5Y%",              "5Y%",            "0.00%",           8),
    # ── Fundamentals ──
    ("V",  "Country",          "country",        "@",              14),
    ("W",  "Sector",           "sector",         "@",              18),
    ("X",  "Industry",         "industry",       "@",              22),
    ("Y",  "Market Cap",       "market_cap",     '#,##0,,"M"',     14),
    ("Z",  "P/E (TTM)",        "pe_ratio",       "0.00",           10),
    ("AA", "52W High",         "52w_high",       "#,##0.00",       12),
    ("AB", "52W Low",          "52w_low",        "#,##0.00",       12),
    ("AC", "% from 52W High",  None,             "0.00%",          15),
    ("AD", "Div. Yield %",     "dividend_yield", "0.00%",          12),
    ("AE", "Beta",             "beta",           "0.00",            8),
    # ── EUR normalisation ──
    ("AF", "FX Rate\n(->EUR)", "fx_rate_eur",    "0.0000",         12),
    ("AG", "Total Value\n(EUR)", None,            "#,##0.00",       15),
    ("AH", "P&L\n(EUR)",       None,             "#,##0.00",       13),
]

# Returns columns L–U (indices 11–20)
PCT_COLS = COLUMNS[11:21]

# Last column letter (for title merge / autofilter)
LAST_COL = COLUMNS[-1][0]   # "AH"

# Valid ticker pattern: uppercase letters, digits, dots, hyphens, carets, equals
TICKER_RE = re.compile(r"^[A-Z0-9.\-\^=]+$")


# ─── Utilities ────────────────────────────────────────────────────────────────

def log(msg):
    """Print a message that doesn't corrupt a tqdm progress bar."""
    if HAS_TQDM:
        _tqdm.write(msg)
    else:
        print(msg)


# ─── Style helpers ────────────────────────────────────────────────────────────

def mk_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def mk_font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, italic=italic, color=color, name="Calibri", size=size)


def mk_center():
    return Alignment(horizontal="center", vertical="center")


def mk_vcenter():
    return Alignment(vertical="center")


def mk_wrap():
    return Alignment(wrap_text=True, vertical="top")


# ─── Data fetching ────────────────────────────────────────────────────────────

def validate_tickers(entries):
    """Warn early about obviously malformed ticker symbols."""
    for entry in entries:
        sym = entry["ticker"]
        if not TICKER_RE.match(sym):
            print(f"  Warning: '{sym}' contains unexpected characters — "
                  "it may not be a valid Yahoo Finance symbol.")


def read_tickers(filepath):
    """Parse tickers file. Returns list of dicts: {ticker, shares, buy_date}.

    Supports inline comments: anything after '#' on a data line is stripped.
    """
    entries = []
    with open(filepath, encoding="utf-8") as f:
        for line in f:
            # Strip inline comments
            if "#" in line:
                line = line[:line.index("#")]
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]
            ticker = parts[0].upper()
            if not ticker:
                continue

            shares = None
            if len(parts) > 1 and parts[1]:
                try:
                    shares = float(parts[1])
                except ValueError:
                    print(f"  Warning: invalid share count '{parts[1]}' for {ticker}, ignoring.")

            buy_date = None
            if len(parts) > 2 and parts[2]:
                try:
                    datetime.strptime(parts[2], "%Y-%m-%d")
                    buy_date = parts[2]
                except ValueError:
                    print(f"  Warning: invalid date '{parts[2]}' for {ticker}, ignoring.")

            entries.append({"ticker": ticker, "shares": shares, "buy_date": buy_date})
    return entries


def normalize_history(hist):
    """Strip timezone so that date arithmetic works simply."""
    if hist.empty:
        return hist
    if getattr(hist.index, "tz", None) is not None:
        hist = hist.copy()
        hist.index = hist.index.tz_convert("UTC").tz_localize(None)
    return hist


def pct_change(hist, days):
    """Percentage change from `days` calendar days ago to the latest close."""
    if hist.empty:
        return None
    latest_price = float(hist["Close"].iloc[-1])
    target_date  = hist.index[-1] - pd.Timedelta(days=days)
    past = hist[hist.index <= target_date]
    if past.empty:
        return None
    past_price = float(past["Close"].iloc[-1])
    if not past_price or pd.isna(past_price) or past_price == 0:
        return None
    return (latest_price - past_price) / past_price


def ytd_change(hist):
    """Percentage change from last trading day of the previous year."""
    if hist.empty:
        return None
    latest_price  = float(hist["Close"].iloc[-1])
    prev_year_end = pd.Timestamp(hist.index[-1].year - 1, 12, 31)
    past = hist[hist.index <= prev_year_end]
    if past.empty:
        return None
    past_price = float(past["Close"].iloc[-1])
    if not past_price or pd.isna(past_price) or past_price == 0:
        return None
    return (latest_price - past_price) / past_price


def fetch_stock(entry):
    """Fetch all data for one ticker entry. Returns dict with 'error' key on failure."""
    symbol   = entry["ticker"]
    buy_date = entry.get("buy_date")

    result = {"ticker": symbol, "input_shares": entry.get("shares"), "buy_date": buy_date}

    try:
        ticker   = yf.Ticker(symbol)
        hist_raw = ticker.history(period="6y")
        try:
            info = ticker.info or {}
        except Exception:
            info = {}
    except Exception as exc:
        result["error"] = str(exc)
        return result

    hist = normalize_history(hist_raw)
    if hist.empty:
        result["error"] = "No history data returned"
        return result

    current_price = float(hist["Close"].iloc[-1])

    timeframes = {
        "1W%": 7, "1M%": 30, "3M%": 91,
        "6M%": 182, "1Y%": 365, "2Y%": 730,
        "3Y%": 1095, "4Y%": 1460, "5Y%": 1825,
    }
    for key, days in timeframes.items():
        result[key] = pct_change(hist, days)
    result["YTD%"] = ytd_change(hist)

    cutoff  = hist.index[-1] - pd.Timedelta(days=365)
    hist_1y = hist[hist.index >= cutoff]
    result["52w_high"] = float(hist_1y["High"].max()) if not hist_1y.empty else None
    result["52w_low"]  = float(hist_1y["Low"].min())  if not hist_1y.empty else None

    result["buy_price"] = None
    if buy_date:
        buy_ts      = pd.Timestamp(buy_date)
        on_or_after = hist[hist.index >= buy_ts]
        if not on_or_after.empty:
            result["buy_price"] = float(on_or_after["Close"].iloc[0])
            actual = on_or_after.index[0].strftime("%Y-%m-%d")
            if actual != buy_date:
                log(f"    i {symbol}: {buy_date} not a trading day, using {actual}")
        else:
            log(f"    ! {symbol}: buy date {buy_date} is beyond available history")

    result.update({
        "name":           info.get("longName") or info.get("shortName") or symbol,
        "currency":       info.get("currency", "USD"),
        "country":        info.get("country")  or "",
        "sector":         info.get("sector")   or "",
        "industry":       info.get("industry") or "",
        "market_cap":     info.get("marketCap"),
        "pe_ratio":       info.get("trailingPE"),
        "beta":           info.get("beta"),
        "dividend_yield": info.get("dividendYield"),
        "current_price":  current_price,
    })

    return result


def fetch_fx_rates(currencies):
    """Fetch EUR conversion rates for a set of currency codes via yfinance.

    Returns dict {currency_code: rate_to_eur}. EUR itself is always 1.0.
    Falls back to 1.0 with a warning if a rate cannot be fetched.
    """
    rates = {"EUR": 1.0}
    non_eur = sorted(c for c in currencies if c != "EUR")
    if not non_eur:
        return rates
    print(f"  Fetching FX rates: {', '.join(non_eur)} -> EUR")
    for cur in non_eur:
        symbol = f"{cur}EUR=X"
        try:
            hist = normalize_history(yf.Ticker(symbol).history(period="5d"))
            if not hist.empty:
                rates[cur] = float(hist["Close"].iloc[-1])
                print(f"    + {cur}/EUR = {rates[cur]:.4f}")
            else:
                print(f"    ! {cur}/EUR: no data returned, defaulting to 1.0")
                rates[cur] = 1.0
        except Exception as exc:
            print(f"    ! {cur}/EUR: {exc}, defaulting to 1.0")
            rates[cur] = 1.0
    return rates


# ─── Sheet: Portfolio ─────────────────────────────────────────────────────────

def create_portfolio_sheet(wb, stocks):
    ws = wb.active
    ws.title = "Portfolio"

    # ── Row 1: title bar ──
    ws.merge_cells(f"A1:{LAST_COL}1")
    c = ws["A1"]
    c.value     = f"Stock Portfolio  ·  Generated {datetime.now():%Y-%m-%d %H:%M}"
    c.font      = mk_font(bold=True, color="FFFFFF", size=13)
    c.fill      = mk_fill(DARK_BLUE)
    c.alignment = mk_center()
    ws.row_dimensions[1].height = 26

    # ── Row 2: thin spacer ──
    ws.row_dimensions[2].height = 4

    # ── Row 3: column headers ──
    for col_letter, header, *_ in COLUMNS:
        c = ws[f"{col_letter}3"]
        c.value     = header
        c.font      = mk_font(bold=True, color="FFFFFF", size=10)
        c.fill      = mk_fill(MED_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 32

    # ── Data rows (starting row 4) ──
    n = len(stocks)
    for i, stock in enumerate(stocks):
        row       = i + 4
        row_fill  = mk_fill(ALT_ROW) if i % 2 else None
        base_font = mk_font()

        def _c(col):
            cell = ws[f"{col}{row}"]
            if row_fill:
                cell.fill = row_fill
            cell.font = base_font
            return cell

        # ── Identity ──
        c = _c("A");  c.value = stock["ticker"];              c.alignment = mk_center()
        c = _c("B");  c.value = stock.get("name", "");        c.alignment = mk_vcenter()
        c = _c("C");  c.value = stock.get("current_price");   c.number_format = "#,##0.00";   c.alignment = mk_center()
        c = _c("D");  c.value = stock.get("currency", "");    c.alignment = mk_center()

        # ── E: Buy Date ──
        c = ws[f"E{row}"]
        if row_fill: c.fill = row_fill
        c.font = base_font;  c.alignment = mk_center()
        if stock.get("buy_date"):
            try:
                c.value = datetime.strptime(stock["buy_date"], "%Y-%m-%d").date()
                c.number_format = "yyyy-mm-dd"
            except ValueError:
                c.value = stock["buy_date"]

        # ── F: Buy Price ──
        c = ws[f"F{row}"]
        if row_fill: c.fill = row_fill
        c.font = base_font;  c.alignment = mk_center()
        if stock.get("buy_price") is not None:
            c.value = stock["buy_price"];  c.number_format = "#,##0.00"

        # ── G: # Shares (user input, yellow; pre-filled if in tickers file) ──
        c = ws[f"G{row}"]
        c.fill          = mk_fill(YELLOW_FILL)
        c.font          = base_font
        c.alignment     = mk_center()
        c.number_format = "#,##0.####"
        if stock.get("input_shares") is not None:
            c.value = stock["input_shares"]

        # ── H: Total Value (formula, green) ──
        c = ws[f"H{row}"]
        c.value         = f'=IF(OR(G{row}="",G{row}=0),"",C{row}*G{row})'
        c.number_format = "#,##0.00"
        c.fill          = mk_fill(GREEN_FILL)
        c.font          = base_font;  c.alignment = mk_center()

        # ── I: Cost Basis (formula, green) ──
        c = ws[f"I{row}"]
        c.value         = f'=IF(OR(F{row}="",G{row}="",G{row}=0),"",F{row}*G{row})'
        c.number_format = "#,##0.00"
        c.fill          = mk_fill(GREEN_FILL)
        c.font          = base_font;  c.alignment = mk_center()

        # ── J: P&L $ (formula) ──
        c = ws[f"J{row}"]
        c.value         = f'=IF(OR(H{row}="",I{row}=""),"",H{row}-I{row})'
        c.number_format = "#,##0.00"
        if row_fill: c.fill = row_fill
        c.font = base_font;  c.alignment = mk_center()

        # ── K: P&L % (formula) ──
        c = ws[f"K{row}"]
        c.value         = f'=IF(OR(F{row}="",F{row}=0),"",(C{row}-F{row})/F{row})'
        c.number_format = "0.00%"
        if row_fill: c.fill = row_fill
        c.font = base_font;  c.alignment = mk_center()

        # ── L–U: % changes ──
        for col_letter, _, key, fmt, _ in PCT_COLS:
            c = _c(col_letter)
            c.value = stock.get(key);  c.number_format = fmt;  c.alignment = mk_center()

        # ── Fundamentals ──
        c = _c("V");  c.value = stock.get("country", "");   c.alignment = mk_vcenter()
        c = _c("W");  c.value = stock.get("sector", "");    c.alignment = mk_vcenter()
        c = _c("X");  c.value = stock.get("industry", "");  c.alignment = mk_vcenter()
        c = _c("Y");  c.value = stock.get("market_cap");    c.number_format = '#,##0,,"M"'; c.alignment = mk_center()
        c = _c("Z");  c.value = stock.get("pe_ratio");      c.number_format = "0.00";       c.alignment = mk_center()
        c = _c("AA"); c.value = stock.get("52w_high");      c.number_format = "#,##0.00";   c.alignment = mk_center()
        c = _c("AB"); c.value = stock.get("52w_low");       c.number_format = "#,##0.00";   c.alignment = mk_center()

        # ── AC: % from 52W High (formula) ──
        c = ws[f"AC{row}"]
        c.value         = f'=IF(OR(AA{row}="",AA{row}=0),"",(C{row}-AA{row})/AA{row})'
        c.number_format = "0.00%"
        if row_fill: c.fill = row_fill
        c.font = base_font;  c.alignment = mk_center()

        c = _c("AD"); c.value = stock.get("dividend_yield"); c.number_format = "0.00%"; c.alignment = mk_center()
        c = _c("AE"); c.value = stock.get("beta");           c.number_format = "0.00";  c.alignment = mk_center()

        # ── AF: FX Rate (->EUR) ──
        c = _c("AF")
        c.value         = stock.get("fx_rate_eur", 1.0)
        c.number_format = "0.0000"
        c.alignment     = mk_center()

        # ── AG: Total Value EUR (formula, green) ──
        c = ws[f"AG{row}"]
        c.value         = f'=IF(OR(H{row}=""),"",H{row}*AF{row})'
        c.number_format = "#,##0.00"
        c.fill          = mk_fill(GREEN_FILL)
        c.font          = base_font;  c.alignment = mk_center()

        # ── AH: P&L EUR (formula) ──
        c = ws[f"AH{row}"]
        c.value         = f'=IF(OR(J{row}=""),"",J{row}*AF{row})'
        c.number_format = "#,##0.00"
        if row_fill: c.fill = row_fill
        c.font = base_font;  c.alignment = mk_center()

    # ── TOTAL row ──
    total_row = n + 4

    c = ws[f"A{total_row}"]
    c.value = "TOTAL";  c.font = mk_font(bold=True, color="FFFFFF", size=11)
    c.fill = mk_fill(DARK_BLUE);  c.alignment = mk_center()

    for col in ("H", "I", "J", "AG", "AH"):
        c = ws[f"{col}{total_row}"]
        c.value         = f"=SUM({col}4:{col}{total_row - 1})"
        c.number_format = "#,##0.00"
        c.font          = mk_font(bold=True, color="FFFFFF", size=11)
        c.fill          = mk_fill(GREEN_TOTAL)
        c.alignment     = mk_center()

    # ── Conditional formatting ──
    data_end = total_row - 1
    red_fill_cf   = PatternFill(start_color=RED_CF,   end_color=RED_CF,   fill_type="solid")
    green_fill_cf = PatternFill(start_color=GREEN_CF, end_color=GREEN_CF, fill_type="solid")

    for cf_range in [f"J4:K{data_end}", f"L4:U{data_end}", f"AH4:AH{data_end}"]:
        ws.conditional_formatting.add(cf_range, CellIsRule(
            operator="lessThan",    formula=["0"], fill=red_fill_cf,   font=Font(color=RED_CF_FONT)))
        ws.conditional_formatting.add(cf_range, CellIsRule(
            operator="greaterThan", formula=["0"], fill=green_fill_cf, font=Font(color=GREEN_CF_FONT)))

    # ── Freeze panes: columns A–D + rows 1–3 ──
    ws.freeze_panes = "E4"

    # ── AutoFilter ──
    ws.auto_filter.ref = f"A3:{LAST_COL}{data_end}"

    # ── Column widths ──
    for col_letter, _, _, _, width in COLUMNS:
        ws.column_dimensions[col_letter].width = width


# ─── Sheet: Distribution ──────────────────────────────────────────────────────

def create_distribution_sheet(wb, stocks):
    ws = wb.create_sheet("Distribution")
    n = len(stocks)

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Portfolio Distribution";  c.font = mk_font(bold=True, color="FFFFFF", size=13)
    c.fill = mk_fill(DARK_BLUE);  c.alignment = mk_center()
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 4

    for col, label in [
        ("A", "Country"), ("B", "# Stocks"), ("C", "% Count"), ("D", "Total Value (EUR)"),
        ("F", "Sector"),  ("G", "# Stocks"), ("H", "% Count"), ("I", "Total Value (EUR)"),
    ]:
        c = ws[f"{col}3"]
        c.value = label;  c.font = mk_font(bold=True, color="FFFFFF", size=10)
        c.fill = mk_fill(MED_BLUE);  c.alignment = mk_center()
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 4

    countries = sorted({s.get("country", "") for s in stocks if s.get("country")})
    sectors   = sorted({s.get("sector",  "") for s in stocks if s.get("sector")})

    for i, country in enumerate(countries):
        row = i + 5;  rf = mk_fill(ALT_ROW) if i % 2 else None
        c = ws[f"A{row}"];  c.value = country;  c.font = mk_font()
        if rf: c.fill = rf
        for col, formula, fmt in [
            ("B", f"=COUNTIF(Portfolio!V:V,A{row})",               "0"),
            ("C", f'=IF(B{row}=0,"",B{row}/{n})',                  "0.00%"),
            ("D", f"=SUMIF(Portfolio!V:V,A{row},Portfolio!AG:AG)", "#,##0.00"),
        ]:
            c = ws[f"{col}{row}"];  c.value = formula;  c.number_format = fmt
            c.alignment = mk_center();  c.font = mk_font()
            if rf: c.fill = rf

    for i, sector in enumerate(sectors):
        row = i + 5;  rf = mk_fill(ALT_ROW) if i % 2 else None
        c = ws[f"F{row}"];  c.value = sector;  c.font = mk_font()
        if rf: c.fill = rf
        for col, formula, fmt in [
            ("G", f"=COUNTIF(Portfolio!W:W,F{row})",               "0"),
            ("H", f'=IF(G{row}=0,"",G{row}/{n})',                  "0.00%"),
            ("I", f"=SUMIF(Portfolio!W:W,F{row},Portfolio!AG:AG)", "#,##0.00"),
        ]:
            c = ws[f"{col}{row}"];  c.value = formula;  c.number_format = fmt
            c.alignment = mk_center();  c.font = mk_font()
            if rf: c.fill = rf

    for col, width in [
        ("A", 18), ("B", 10), ("C", 10), ("D", 14),
        ("E",  3),
        ("F", 22), ("G", 10), ("H", 10), ("I", 14),
    ]:
        ws.column_dimensions[col].width = width


# ─── Sheet: Charts ────────────────────────────────────────────────────────────

def create_charts_sheet(wb, stocks):
    try:
        from openpyxl.chart import PieChart, BarChart, Reference
    except ImportError:
        return

    ws = wb.create_sheet("Charts")

    # ── Title ──
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "Portfolio Charts"
    c.font      = mk_font(bold=True, color="FFFFFF", size=13)
    c.fill      = mk_fill(DARK_BLUE)
    c.alignment = mk_center()
    ws.row_dimensions[1].height = 26

    # ── Sector distribution data table (cols A-B, starting row 2) ──
    sectors = {}
    for s in stocks:
        sec = s.get("sector") or "Unknown"
        sectors[sec] = sectors.get(sec, 0) + 1

    for col, label in [("A", "Sector"), ("B", "Count")]:
        c = ws[f"{col}2"]
        c.value     = label
        c.font      = mk_font(bold=True, color="FFFFFF")
        c.fill      = mk_fill(MED_BLUE)
        c.alignment = mk_center()
    ws.row_dimensions[2].height = 22

    sector_list = sorted(sectors.items(), key=lambda x: -x[1])
    for i, (sec, cnt) in enumerate(sector_list):
        r = i + 3
        c = ws.cell(row=r, column=1, value=sec)
        c.font = mk_font()
        if i % 2:
            c.fill = mk_fill(ALT_ROW)
        c = ws.cell(row=r, column=2, value=cnt)
        c.font      = mk_font()
        c.alignment = mk_center()
        if i % 2:
            c.fill = mk_fill(ALT_ROW)

    sector_end = 2 + len(sector_list)

    if sector_list:
        pie = PieChart()
        pie.title  = "Sector Distribution"
        pie.style  = 10
        pie.width  = 16
        pie.height = 12
        labels = Reference(ws, min_col=1, min_row=3, max_row=sector_end)
        data   = Reference(ws, min_col=2, min_row=2, max_row=sector_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, "D2")

    # ── P&L % bar chart — only stocks with buy prices ──
    pnl_items = [
        (s["ticker"], (s["current_price"] - s["buy_price"]) / s["buy_price"])
        for s in stocks
        if s.get("buy_price") and s.get("current_price")
    ]
    pnl_items.sort(key=lambda x: x[1])  # worst to best (left to right)

    if pnl_items:
        pnl_start = sector_end + 3

        for col, label in [("A", "Ticker"), ("B", "P&L %")]:
            c = ws.cell(row=pnl_start, column=1 if col == "A" else 2, value=label)
            c.font      = mk_font(bold=True, color="FFFFFF")
            c.fill      = mk_fill(MED_BLUE)
            c.alignment = mk_center()

        for i, (ticker, pnl) in enumerate(pnl_items):
            r = pnl_start + 1 + i
            ws.cell(row=r, column=1, value=ticker).font = mk_font()
            c = ws.cell(row=r, column=2, value=round(pnl, 6))
            c.number_format = "0.00%"
            c.font          = mk_font()

        pnl_end = pnl_start + len(pnl_items)

        bar = BarChart()
        bar.type          = "col"
        bar.title         = "Unrealised P&L % by Stock"
        bar.style         = 10
        bar.y_axis.title  = "Return"
        bar.x_axis.title  = "Ticker"
        bar.y_axis.numFmt = "0%"
        bar.width         = 20
        bar.height        = 12

        data_ref = Reference(ws, min_col=2, min_row=pnl_start, max_row=pnl_end)
        cats_ref = Reference(ws, min_col=1, min_row=pnl_start + 1, max_row=pnl_end)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws.add_chart(bar, f"D{pnl_start}")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12


# ─── Sheet: Errors ────────────────────────────────────────────────────────────

def create_errors_sheet(wb, failed):
    ws = wb.create_sheet("Errors")

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = f"Failed Tickers  ({len(failed)} total)"
    c.font      = mk_font(bold=True, color="FFFFFF", size=13)
    c.fill      = mk_fill(DARK_RED)
    c.alignment = mk_center()
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 4

    for col, label in [("A", "Ticker"), ("B", "Error"), ("C", "Shares"), ("D", "Buy Date")]:
        c = ws[f"{col}3"]
        c.value     = label
        c.font      = mk_font(bold=True, color="FFFFFF", size=10)
        c.fill      = mk_fill(MED_BLUE)
        c.alignment = mk_center()
    ws.row_dimensions[3].height = 24

    for i, s in enumerate(failed):
        row = i + 4
        rf  = mk_fill(ALT_ROW) if i % 2 else None
        for col, val in [
            ("A", s["ticker"]),
            ("B", s.get("error", "Unknown error")),
            ("C", s.get("input_shares")),
            ("D", s.get("buy_date")),
        ]:
            c = ws[f"{col}{row}"]
            c.value     = val
            c.font      = mk_font()
            c.alignment = mk_vcenter()
            if rf:
                c.fill = rf

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12


# ─── Sheet: AI Analysis ───────────────────────────────────────────────────────

def create_ai_sheet(wb, stocks, api_key=None, auth_token=None, model="claude-sonnet-4-6"):
    import anthropic

    ws = wb.create_sheet("AI Analysis")
    ws.column_dimensions["A"].width = 110

    c = ws["A1"]
    c.value     = f"AI Portfolio Analysis  ·  Powered by Claude ({model})"
    c.font      = mk_font(bold=True, color="FFFFFF", size=13)
    c.fill      = mk_fill(DARK_BLUE)
    c.alignment = mk_center()
    ws.row_dimensions[1].height = 26

    def safe_pct(val):
        return round(val * 100, 2) if val is not None else None

    portfolio_data = [
        {
            "ticker":        s.get("ticker"),
            "name":          s.get("name"),
            "sector":        s.get("sector"),
            "country":       s.get("country"),
            "currency":      s.get("currency"),
            "price":         round(s["current_price"], 2) if s.get("current_price") else None,
            "market_cap_M":  round(s["market_cap"] / 1e6) if s.get("market_cap") else None,
            "pe_ratio":      round(s["pe_ratio"], 2) if s.get("pe_ratio") else None,
            "beta":          round(s["beta"], 2) if s.get("beta") else None,
            "div_yield_pct": safe_pct(s.get("dividend_yield")),
            "ytd_pct":       safe_pct(s.get("YTD%")),
            "1y_pct":        safe_pct(s.get("1Y%")),
            "3y_pct":        safe_pct(s.get("3Y%")),
            "5y_pct":        safe_pct(s.get("5Y%")),
            "buy_date":      s.get("buy_date"),
            "buy_price":     round(s["buy_price"], 2) if s.get("buy_price") else None,
            "pnl_pct": (
                safe_pct((s["current_price"] - s["buy_price"]) / s["buy_price"])
                if s.get("buy_price") and s.get("current_price") else None
            ),
            "pct_from_52w_high": (
                round((s["current_price"] - s["52w_high"]) / s["52w_high"] * 100, 2)
                if s.get("current_price") and s.get("52w_high") else None
            ),
        }
        for s in stocks
    ]

    prompt = (
        "You are a professional financial analyst. Analyse the following stock portfolio "
        "and provide a structured, data-driven report in Markdown.\n\n"
        f"Portfolio data (JSON):\n{json.dumps(portfolio_data, indent=2)}\n\n"
        "Structure your response with these exact section headers:\n\n"
        "## Portfolio Overview\n"
        "Brief summary: number of stocks, currencies, geographic spread, sector mix.\n\n"
        "## Geographic & Sector Diversification\n"
        "Assess diversification. Flag any single geography or sector exceeding ~40%.\n\n"
        "## Performance Highlights\n"
        "Notable outperformers and underperformers by YTD, 1Y, and 5Y. "
        "Where buy dates are available, include unrealised P&L context.\n\n"
        "## Risk Assessment\n"
        "Beta analysis, valuation concerns (elevated P/E), drawdown risks, "
        "and any stocks significantly below their 52-week high.\n\n"
        "## Sector Concentration Warnings\n"
        "Flag any sector representing more than 35% of identifiable holdings.\n\n"
        "## Key Takeaways\n"
        "3-5 concise bullet points of the most important insights.\n\n"
        "Keep the tone professional. Reference specific tickers and numbers. "
        "Append a disclaimer that this is not financial advice."
    )

    print(f"  Calling {model} (streaming)...")
    print("  " + "-" * 56)
    try:
        client_ai = anthropic.Anthropic(auth_token=auth_token) if auth_token else anthropic.Anthropic(api_key=api_key)
        analysis_parts = []
        with client_ai.messages.stream(
            model=model,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            for text in stream.text_stream:
                print(text, end="", flush=True)
                analysis_parts.append(text)
        print()
        print("  " + "-" * 56)
        analysis = "".join(analysis_parts)
        print("  Analysis complete.")
    except Exception as exc:
        analysis = f"Error generating AI analysis: {exc}\n\nCheck that ANTHROPIC_API_KEY is valid."
        print(f"\n  Claude API error: {exc}")

    row = 3
    for line in analysis.splitlines():
        stripped = line.strip()
        c = ws.cell(row=row, column=1)
        c.alignment = mk_wrap()
        if stripped.startswith("## "):
            c.value = stripped[3:];  c.font = mk_font(bold=True, color=DARK_BLUE, size=12)
            c.fill = mk_fill(LIGHT_BLUE);  ws.row_dimensions[row].height = 20
        elif stripped.startswith("# "):
            c.value = stripped[2:];  c.font = mk_font(bold=True, color="FFFFFF", size=13)
            c.fill = mk_fill(DARK_BLUE);   ws.row_dimensions[row].height = 22
        elif stripped == "":
            ws.row_dimensions[row].height = 8
        else:
            c.value = line;  c.font = mk_font(size=10);  ws.row_dimensions[row].height = 15
        row += 1

    row += 1
    c = ws.cell(row=row, column=1)
    c.value = (f"Generated by Claude ({model}) · {datetime.now():%Y-%m-%d %H:%M} · "
               "For informational purposes only — not financial advice.")
    c.font = mk_font(italic=True, color="888888", size=9)


# ─── File opener ──────────────────────────────────────────────────────────────

def open_file(path):
    """Open a file with the system default application."""
    try:
        if sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
        elif sys.platform == "win32":
            os.startfile(str(path))
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
    except Exception as exc:
        print(f"  Could not open file automatically: {exc}")


# ─── Analysis runner ──────────────────────────────────────────────────────────

def run_analysis(args):
    """Run the full analysis pipeline. Returns output Path on success, None on failure."""
    tickers_path = Path(args.tickers_file)
    if not tickers_path.exists():
        print(f"Error: '{tickers_path}' not found.")
        return None

    print(f"\nReading tickers from '{tickers_path}'...")
    entries = read_tickers(str(tickers_path))
    if not entries:
        print("No tickers found. Check your tickers file.")
        return None

    print(f"Found {len(entries)} ticker(s): {', '.join(e['ticker'] for e in entries)}")
    validate_tickers(entries)

    print("\nFetching market data from Yahoo Finance...")
    all_data = []
    if HAS_TQDM:
        iterator = _tqdm(entries, desc="  Fetching", unit="ticker", ncols=72)
    else:
        iterator = entries

    for entry in iterator:
        result = fetch_stock(entry)
        all_data.append(result)
        if "error" in result:
            log(f"  ! {result['ticker']}: {result['error']}")
        else:
            bp = result.get("buy_price")
            bp_str = f"  buy@{bp:,.2f}" if bp else ""
            log(f"  + {result['ticker']}: {result.get('name', '')}  "
                f"{result.get('current_price', 0):,.2f} {result.get('currency', '')}{bp_str}")

    stocks = [s for s in all_data if "error" not in s]
    failed = [s for s in all_data if "error" in s]

    if failed:
        print(f"\nWarning: failed to fetch {len(failed)} ticker(s): "
              f"{', '.join(s['ticker'] for s in failed)}")
    if not stocks:
        print("No valid stock data retrieved.")
        return None

    has_pnl = sum(1 for s in stocks if s.get("buy_price"))
    print(f"\nSuccessfully fetched {len(stocks)} / {len(entries)} stock(s)."
          + (f"  {has_pnl} with P&L data." if has_pnl else ""))

    print("\nFetching FX rates...")
    currencies = {s.get("currency", "USD") for s in stocks}
    fx_rates = fetch_fx_rates(currencies)
    for stock in stocks:
        stock["fx_rate_eur"] = fx_rates.get(stock.get("currency", "USD"), 1.0)

    print("\nBuilding Excel workbook...")
    wb = Workbook()

    create_portfolio_sheet(wb, stocks)
    print("  + Portfolio sheet")

    create_distribution_sheet(wb, stocks)
    print("  + Distribution sheet")

    create_charts_sheet(wb, stocks)
    print("  + Charts sheet")

    if failed:
        create_errors_sheet(wb, failed)
        print(f"  + Errors sheet  ({len(failed)} failed ticker(s))")

    api_key    = os.environ.get("ANTHROPIC_API_KEY")
    auth_token = os.environ.get("ANTHROPIC_AUTH_TOKEN")
    skip_ai    = getattr(args, "no_ai", False)
    model      = getattr(args, "model", "claude-sonnet-4-6")

    if (api_key or auth_token) and not skip_ai:
        print(f"\nGenerating AI analysis (this may take 15-30 s)...")
        create_ai_sheet(wb, stocks, api_key=api_key, auth_token=auth_token, model=model)
        print("  + AI Analysis sheet")
    elif skip_ai:
        print("\n  --no-ai flag set — AI Analysis sheet skipped.")
    else:
        print(
            "\n  Note: ANTHROPIC_API_KEY not set — AI Analysis sheet skipped.\n"
            "  Copy .env.example to .env, add your key, and re-run to enable it."
        )

    # Determine output path
    if getattr(args, "output", None):
        output = Path(args.output)
        output.parent.mkdir(parents=True, exist_ok=True)
    else:
        out_dir = tickers_path.parent.resolve()
        output  = out_dir / f"portfolio_{datetime.now():%Y%m%d_%H%M}.xlsx"

    wb.save(output)

    print(f"\n{'=' * 60}")
    print(f"  Saved: {output}")
    print(f"{'=' * 60}")
    print("\nNext steps:")
    print("  1. Open the file in Excel / LibreOffice Calc")
    print("  2. Adjust '# Shares' (column G, yellow) if needed")
    print("  3. P&L and Distribution sheet update automatically")

    if getattr(args, "auto_open", False):
        open_file(output)

    return output


# ─── Watch mode ───────────────────────────────────────────────────────────────

def watch_mode(tickers_path, args):
    """Poll tickers file and re-run analysis whenever it changes."""
    print(f"Watching '{tickers_path}' for changes  (Ctrl+C to stop)...")
    last_mtime = 0
    first_run  = True
    try:
        while True:
            try:
                mtime = tickers_path.stat().st_mtime
            except FileNotFoundError:
                time.sleep(1)
                continue

            if mtime != last_mtime:
                last_mtime = mtime
                if not first_run:
                    print(f"\n[{datetime.now():%H:%M:%S}] Change detected — regenerating...")
                first_run = False
                run_analysis(args)

            time.sleep(1)
    except KeyboardInterrupt:
        print("\nWatch mode stopped.")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Stock Portfolio Analyzer — yfinance + openpyxl + Claude",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python stock_analyzer.py\n"
            "  python stock_analyzer.py my_stocks.txt --open\n"
            "  python stock_analyzer.py --no-ai --output report.xlsx\n"
            "  python stock_analyzer.py --watch\n"
            "  python stock_analyzer.py --add 'TSLA,10,2024-03-01'\n"
            "  python stock_analyzer.py --model claude-opus-4-6\n"
        ),
    )
    parser.add_argument(
        "tickers_file", nargs="?", default="tickers.txt",
        help="Path to tickers file (default: tickers.txt)",
    )
    parser.add_argument(
        "--output", "-o", metavar="FILE",
        help="Output .xlsx path (default: portfolio_YYYYMMDD_HHMM.xlsx next to tickers file)",
    )
    parser.add_argument(
        "--no-ai", action="store_true", dest="no_ai",
        help="Skip AI analysis even if ANTHROPIC_API_KEY is set",
    )
    parser.add_argument(
        "--open", action="store_true", dest="auto_open",
        help="Open the generated file automatically after saving",
    )
    parser.add_argument(
        "--watch", action="store_true",
        help="Watch tickers file for changes and regenerate automatically",
    )
    parser.add_argument(
        "--model", default="claude-sonnet-4-6", metavar="MODEL",
        help="Claude model for AI analysis (default: claude-sonnet-4-6)",
    )
    parser.add_argument(
        "--add", metavar="ENTRY",
        help="Append a ticker line to tickers file and exit  e.g. 'TSLA,10,2024-03-01'",
    )
    args = parser.parse_args()
    tickers_path = Path(args.tickers_file)

    # ── --add: append ticker and exit ──
    if args.add:
        with open(tickers_path, "a", encoding="utf-8") as f:
            f.write(f"\n{args.add}\n")
        print(f"Added to '{tickers_path}': {args.add}")
        sys.exit(0)

    if not tickers_path.exists():
        print(f"Error: '{tickers_path}' not found.")
        sys.exit(1)

    if args.watch:
        watch_mode(tickers_path, args)
    else:
        result = run_analysis(args)
        if result is None:
            sys.exit(1)


if __name__ == "__main__":
    main()
